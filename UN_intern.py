import logging
import time
import random
import traceback
import re
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
)
from selenium_stealth import stealth
import openpyxl
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
from concurrent.futures import ThreadPoolExecutor, as_completed
import pycountry

# ------------------------ Configuration ------------------------

# Get the current directory
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

# Path to ChromeDriver - using a relative path in the same directory as the script
DRIVER_PATH = os.path.join(CURRENT_DIR, "chromedriver.exe")  

# Base URL of the UN Careers website (Internships)
BASE_URL = "https://careers.un.org/jobopening?language=en&data=%257B%2522jle%2522:%255B%255D,%2522jc%2522:%255B%2522INT%2522%255D%257D"

# Base site URL for constructing full job links
BASE_SITE_URL = "https://careers.un.org"

# Excel file names - using relative paths
EXCEL_FILENAME = os.path.join(CURRENT_DIR, "UN_Internships.xlsx")
APPLIED_EXCEL_FILENAME = os.path.join(CURRENT_DIR, "applied_intern.xlsx")

# Create logs directory if it doesn't exist
LOGS_DIR = os.path.join(CURRENT_DIR, "logs")
os.makedirs(LOGS_DIR, exist_ok=True)

# Initialize geolocator
geolocator = Nominatim(user_agent="un_internships_scraper")

# Caching dictionary to store location-country mappings
location_country_cache = {}

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOGS_DIR, "scraping.log")),
        logging.StreamHandler()
    ]
)

# ------------------------ Helper Functions ------------------------

def sanitize_sheet_name(name, existing_names):
    """
    Sanitizes the sheet name by removing invalid characters and ensuring uniqueness.

    Args:
        name (str): Original sheet name.
        existing_names (set): Set of already used sheet names.

    Returns:
        str: Sanitized and unique sheet name.
    """
    # Remove invalid characters
    sanitized = re.sub(r'[\\/*?:\[\]]', '', name)
    # Replace multiple spaces with a single space
    sanitized = re.sub(r'\s+', ' ', sanitized).strip()
    # Truncate to 31 characters
    sanitized = sanitized[:31]
    original_sanitized = sanitized
    counter = 1
    # Ensure uniqueness
    while sanitized in existing_names:
        suffix = f"_{counter}"
        # Ensure total length does not exceed 31
        if len(original_sanitized) + len(suffix) > 31:
            sanitized = original_sanitized[:31 - len(suffix)] + suffix
        else:
            sanitized = original_sanitized + suffix
        counter += 1
    existing_names.add(sanitized)
    return sanitized

def get_standard_country_name(country_name):
    """
    Converts a country name to its standard English name using pycountry.
    Args:
        country_name (str): The original country name.
    Returns:
        str: The standardized English country name, or the original name if not found.
    """
    try:
        country = pycountry.countries.lookup(country_name)
        return country.name
    except LookupError:
        logging.warning(f"Could not standardize country name: {country_name}")
        return country_name

def load_applied_job_ids(filename=APPLIED_EXCEL_FILENAME):
    """
    Loads applied Job IDs from an Excel file.
    Args:
        filename (str): The path to the Excel file containing applied Job IDs.
    Returns:
        set: A set of Job IDs (as strings) that have been applied to.
    """
    job_ids = set()
    try:
        wb = load_workbook(filename=filename)
        sheet = wb.active  # or specify sheet name if needed
        # Find the column index for 'Job ID'
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if 'Job ID' not in header:
            logging.error(f"'Job ID' column not found in {filename}.")
            return job_ids
        job_id_col = header.index('Job ID') + 1  # openpyxl is 1-based
        for row in sheet.iter_rows(min_row=2, min_col=job_id_col, max_col=job_id_col):
            cell = row[0]
            if cell.value:
                job_ids.add(str(cell.value).strip())
        logging.info(f"Loaded {len(job_ids)} applied Job IDs from {filename}.")
    except FileNotFoundError:
        logging.error(f"File {filename} not found.")
    except Exception as e:
        logging.error(f"Error reading {filename}: {e}")
    
    return job_ids

# ------------------------ Selenium Setup ------------------------

def configure_driver():
    """
    Configures and returns a Selenium WebDriver instance with stealth settings.
    """
    chrome_options = Options()
    
    # Uncomment the following line to run in headless mode
    # Note: Headless mode may increase the chances of being detected
    # chrome_options.add_argument("--headless=new")  # Use 'new' for Chrome 109+
    
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    
    # Set a common User-Agent
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/112.0.0.0 Safari/537.36"
    )
    
    # Disable Selenium flags
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    # Initialize WebDriver
    service = Service(DRIVER_PATH)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    # Apply selenium-stealth to minimize detection
    stealth(
        driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
    )
    
    return driver

# ------------------------ Scraping Functions ------------------------

def accept_cookies(driver):
    """
    Accepts cookies if the consent prompt appears.
    """
    try:
        wait = WebDriverWait(driver, 10)
        # Example: Adjust the XPath based on actual HTML
        accept_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Accept Cookies')]"))
        )
        accept_button.click()
        logging.info("Accepted cookies.")
        # Randomized short delay after accepting cookies
        time.sleep(random.uniform(1, 2))
    except (TimeoutException, NoSuchElementException):
        logging.info("No cookie consent prompt found.")

def close_floating_elements(driver):
    """
    Closes any floating elements or overlays that might interfere with clicking actions.
    """
    try:
        # Example: Adjust the XPath based on actual HTML of the floating element
        floating_menu = driver.find_element(By.XPATH, "//div[contains(@class, 'floating-menu-class')]")
        close_button = floating_menu.find_element(By.XPATH, ".//button[@aria-label='Close']")
        close_button.click()
        logging.info("Closed floating menu.")
        # Wait briefly after closing
        time.sleep(random.uniform(1, 2))
    except NoSuchElementException:
        logging.info("No floating menu found to close.")
    except Exception as e:
        logging.error(f"Error closing floating menu: {e}")
        logging.error(traceback.format_exc())

def set_records_per_page(driver, records=50):
    """
    Sets the number of records per page to the specified number using JavaScript to bypass overlay issues.
    """
    try:
        wait = WebDriverWait(driver, 15)
        # Locate the 'Records per Page' dropdown label
        records_label = wait.until(
            EC.presence_of_element_located((By.XPATH, "//label[contains(text(), 'Records per Page:')]"))
        )
        # Locate the corresponding dropdown button
        dropdown_button = records_label.find_element(By.XPATH, "../div//button[contains(@class, 'dropdown-toggle')]")
        dropdown_button.click()
        logging.info("Clicked on 'Records per Page' dropdown.")
        
        # Wait for the dropdown options to appear
        time.sleep(random.uniform(1, 2))
        
        # Locate the desired option
        desired_option = wait.until(
            EC.presence_of_element_located((By.XPATH, f"//button[@class='dropdown-item pt-1' and text()='{records}']"))
        )
        
        # Scroll the desired option into view
        driver.execute_script("arguments[0].scrollIntoView(true);", desired_option)
        
        # Close any floating menus that might be obstructing the click
        close_floating_elements(driver)
        
        # Use JavaScript to click the desired option
        driver.execute_script("arguments[0].click();", desired_option)
        logging.info(f"Set 'Records per Page' to {records} via JavaScript click.")
        
        # Wait for the page to reload after changing records per page
        time.sleep(random.uniform(2, 4))
    except (TimeoutException, NoSuchElementException, ElementClickInterceptedException) as e:
        logging.error(f"Error setting records per page: {e}")
        logging.error(traceback.format_exc())
        # Take a screenshot for debugging
        timestamp = int(time.time())
        screenshot_path = os.path.join(LOGS_DIR, f"set_records_per_page_error_{timestamp}.png")
        driver.save_screenshot(screenshot_path)

def get_job_elements(driver):
    """
    Retrieves all job listing elements from the current page.
    """
    try:
        jobs = driver.find_elements(By.XPATH, "//div[contains(@class, 'card border-0 ng-star-inserted')]")
        logging.info(f"Found {len(jobs)} job elements on the current page.")
        return jobs
    except NoSuchElementException:
        logging.error("Job elements not found.")
        return []

def extract_job_details(job_element):
    """
    Extracts the job details from a job element, including City.
    """
    try:
        # Extract Job Title
        title_element = job_element.find_element(By.XPATH, ".//h2[contains(@class, 'jbOpen_title')]")
        title = title_element.text.strip()
        
        # Extract Job ID
        job_id_element = job_element.find_element(By.XPATH, ".//span[contains(@class, 'jbOpen_Id')]")
        job_id_text = job_id_element.text.strip()
        # Extract numeric Job ID using split
        job_id = job_id_text.split(":")[-1].strip()
        
        # Extract other details from card-body
        card_body = job_element.find_element(By.XPATH, ".//div[contains(@class, 'card-body')]")
        body_text = card_body.text.strip().split('\n')
        
        # Initialize dictionary
        job_details = {
            "Title": title,
            "Job ID": job_id,
            "Job Network": "",
            "Job Family": "",
            "Category and Level": "",
            "Duty Station": "",
            "City": "",
            # Country will be added later
            "Department/Office": "",
            "Date Posted": "",
            "Deadline": "",
            "Job Description Link": ""
        }
        
        # Iterate over each line to extract details
        for line in body_text:
            if "Job Network" in line:
                job_details["Job Network"] = line.split("Job Network :")[-1].strip()
            elif "Job Family" in line:
                job_details["Job Family"] = line.split("Job Family :")[-1].strip()
            elif "Category and Level" in line:
                job_details["Category and Level"] = line.split("Category and Level :")[-1].strip()
            elif "Duty Station" in line:
                duty_station = line.split("Duty Station :")[-1].strip()
                job_details["Duty Station"] = duty_station
                # Extract city from duty station
                if "," in duty_station:
                    location_query = duty_station  # Use entire Duty Station for geocoding
                    city = duty_station.split(",")[0].strip()
                else:
                    location_query = duty_station
                    city = duty_station
                job_details["City"] = city
            elif "Department/Office" in line:
                job_details["Department/Office"] = line.split("Department/Office :")[-1].strip()
            elif "Date Posted" in line:
                job_details["Date Posted"] = line.split("Date Posted :")[-1].strip()
            elif "Deadline" in line:
                # Corrected extraction: Deadline is within the same <span>
                job_details["Deadline"] = line.split("Deadline :")[-1].strip()
        
        # Extract Job Description Link
        try:
            link_element = card_body.find_element(By.XPATH, ".//a[contains(@class, 'btn btn-primary')]")
            relative_link = link_element.get_attribute("href")
            # Construct full URL if relative
            if relative_link.startswith("/"):
                full_link = BASE_SITE_URL + relative_link
            else:
                full_link = relative_link
            job_details["Job Description Link"] = full_link
        except NoSuchElementException:
            logging.warning(f"Job Description link not found for Job ID: {job_id}")
            job_details["Job Description Link"] = ""
        
        return job_details
    except Exception as e:
        logging.error(f"Error extracting job details: {e}")
        logging.error(traceback.format_exc())
        return None

def click_next_page(driver):
    """
    Clicks the 'Next' button to navigate to the next page of job listings.
    Returns True if the next page was clicked, False otherwise.
    """
    try:
        wait = WebDriverWait(driver, 10)
        # Locate the 'Next' button by aria-label
        next_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[@aria-label='Next']"))
        )
        # Check if the parent li has 'disabled' class
        parent_li = next_button.find_element(By.XPATH, "./parent::li")
        classes = parent_li.get_attribute("class")
        if "disabled" in classes:
            logging.info("No more pages to navigate.")
            return False
        else:
            # Scroll to the 'Next' button to ensure it's in view
            driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
            # Click the 'Next' button via JavaScript to bypass overlays
            driver.execute_script("arguments[0].click();", next_button)
            logging.info("Clicked on 'Next' page button.")
            # Wait for the next page to load
            time.sleep(random.uniform(2, 4))
            return True
    except (TimeoutException, NoSuchElementException, ElementClickInterceptedException) as e:
        logging.error(f"Error clicking 'Next' page: {e}")
        logging.error(traceback.format_exc())
        # Take a screenshot for debugging
        timestamp = int(time.time())
        screenshot_path = os.path.join(LOGS_DIR, f"click_next_page_error_{timestamp}.png")
        driver.save_screenshot(screenshot_path)
        return False

def get_internship_data():
    """
    Scrapes internship opportunities from the UN Careers website.
    Returns:
        A list of dictionaries containing internship details.
    """
    internships = []
    driver = configure_driver()
    
    try:
        logging.info(f"Navigating to {BASE_URL}")
        driver.get(BASE_URL)
        
        # Wait for the main content to load
        wait = WebDriverWait(driver, 15)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        logging.info("Page loaded successfully.")
        
        # Accept cookies if prompted
        accept_cookies(driver)
        
        # Set 'Records per Page' to 50 to minimize pagination
        set_records_per_page(driver, records=50)
        
        page_number = 1
        while True:
            logging.info(f"Scraping page {page_number}...")
            # Retrieve job elements
            jobs = get_job_elements(driver)
            
            if not jobs:
                logging.warning("No jobs found on this page. Saving page source for debugging.")
                debug_file = os.path.join(LOGS_DIR, f"page_source_page_{page_number}.html")
                with open(debug_file, "w", encoding="utf-8") as file:
                    file.write(driver.page_source)
                break
            
            for job in jobs:
                details = extract_job_details(job)
                if details:
                    internships.append(details)
                    logging.info(f"Extracted: {details['Title']} - Job ID: {details['Job ID']}")
                # Randomized short delay between processing jobs
                time.sleep(random.uniform(0.2, 0.5))
            
            # Attempt to click 'Next' to go to the next page
            has_next = click_next_page(driver)
            if not has_next:
                break
            page_number += 1
        
    except Exception as e:
        logging.error(f"An error occurred during scraping: {e}")
        logging.error(traceback.format_exc())
    finally:
        driver.quit()
        logging.info("WebDriver closed.")
    
    return internships

# ------------------------ Geocoding Functions ------------------------

def geocode_city(city):
    """
    Geocodes a city to find its country.
    """
    try:
        # Sleep to respect Nominatim's rate limit (1 request per second)
        time.sleep(1)
        location = geolocator.geocode(city, exactly_one=True, timeout=10)
        if location and location.address:
            address = location.address
            # Extract country from the address
            country = address.split(",")[-1].strip()
            # Standardize country name to English
            country = get_standard_country_name(country)
            logging.info(f"Identified country for city '{city}': {country}")
            return city, country
        else:
            logging.warning(f"Could not identify country for city: {city}")
            return city, "Unknown"
    except (GeocoderTimedOut, GeocoderServiceError) as e:
        logging.error(f"Geocoding error for city '{city}': {e}")
        return city, "Unknown"

def geocode_cities(cities):
    """
    Geocodes a list of cities using multithreading.
    """
    city_country_dict = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_city = {executor.submit(geocode_city, city): city for city in cities}
        for future in as_completed(future_to_city):
            city = future_to_city[future]
            try:
                city_result, country = future.result()
                city_country_dict[city_result] = country
            except Exception as e:
                logging.error(f"Error geocoding city '{city}': {e}")
                city_country_dict[city] = "Unknown"
    return city_country_dict

# ------------------------ Excel Handling ------------------------

def save_to_excel(data, filename=EXCEL_FILENAME):
    """
    Saves internship data to an Excel file, organizing them into separate sheets by country.
    Args:
        data: List of dictionaries containing internship details.
        filename: Name of the Excel file to save data.
    """
    if not data:
        logging.warning("No data to save.")
        return
    
    # Collect unique cities
    cities = set(item.get("City", "Unknown") for item in data)
    logging.info(f"Unique cities to geocode: {len(cities)}")
    
    # Geocode cities using multithreading
    city_country_dict = geocode_cities(cities)
    
    # Load applied Job IDs
    applied_job_ids = load_applied_job_ids()
    if applied_job_ids:
        logging.info(f"Filtering out {len(applied_job_ids)} applied internships.")
    else:
        logging.info("No applied internships to filter out.")
    
    # Update data with country information and exclude applied internships
    filtered_data = []
    for item in data:
        job_id = item.get("Job ID", "")
        if job_id in applied_job_ids:
            logging.info(f"Excluding applied internship with Job ID: {job_id}")
            continue  # Skip this internship
        city = item.get("City", "Unknown")
        country = city_country_dict.get(city, "Unknown")
        # Standardize country name to English
        country = get_standard_country_name(country)
        item["Country"] = country
        filtered_data.append(item)
    
    logging.info(f"Total internships after filtering: {len(filtered_data)}")
    
    # Create a new workbook
    workbook = openpyxl.Workbook()
    
    # Organize data by country
    country_dict = {}
    for item in filtered_data:
        country = item.get("Country", "Unknown")
        if country not in country_dict:
            country_dict[country] = []
        country_dict[country].append(item)
    
    # Keep track of existing sheet names to ensure uniqueness
    existing_sheet_names = set()
    
    # Iterate over each country and create a separate sheet
    for idx, (country, jobs) in enumerate(country_dict.items()):
        # Sanitize sheet name
        sanitized_sheet_name = sanitize_sheet_name(country, existing_sheet_names)
        
        if idx == 0:
            # Rename the default sheet for the first country
            sheet = workbook.active
            sheet.title = sanitized_sheet_name
        else:
            # Create a new sheet for other countries
            sheet = workbook.create_sheet(title=sanitized_sheet_name)
        
        # Add headers
        headers = [
            "Title",
            "Job ID",
            "Job Network",
            "Job Family",
            "Category and Level",
            "Duty Station",
            "City",
            "Country",
            "Department/Office",
            "Date Posted",
            "Deadline",
            "Job Description Link"
        ]
        sheet.append(headers)
        logging.info(f"Excel headers added to sheet: {sanitized_sheet_name}")
        
        # Add job rows
        for job in jobs:
            sheet.append([
                job.get("Title", ""),
                job.get("Job ID", ""),
                job.get("Job Network", ""),
                job.get("Job Family", ""),
                job.get("Category and Level", ""),
                job.get("Duty Station", ""),
                job.get("City", ""),
                job.get("Country", ""),
                job.get("Department/Office", ""),
                job.get("Date Posted", ""),
                job.get("Deadline", ""),
                job.get("Job Description Link", "")
            ])
        logging.info(f"Added {len(jobs)} rows to sheet: {sanitized_sheet_name}")
    
    # Save the workbook (overwrites existing file)
    workbook.save(filename)
    logging.info(f"Data saved to {filename} successfully.")

# ------------------------ Main Execution ------------------------

def main():
    logging.info("Scraping process started.")
    internships = get_internship_data()
    
    if internships:
        logging.info(f"Found {len(internships)} internships. Saving to Excel...")
        save_to_excel(internships)
        logging.info("Scraping process completed successfully.")
    else:
        logging.warning("No internships found. Please check the logs for details.")

if __name__ == "__main__":
    main()
